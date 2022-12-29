import multiprocessing
import pandas as pd
import xlsxwriter
import openpyxl

from os import listdir, walk
from os.path import join, isfile
from mainprocess import mainprocess
from infodict import total_df_human_column_list, total_df_zombie_column_list, \
    total_df_human_excel_column_list, total_df_zombie_excel_column_list


pd.set_option('display.max_colwidth', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.width', 2000)
pd.set_option("expand_frame_repr", True)


def splitlist(list, n):
    len_list = len(list)
    len_sublist = len_list//n + 1
    result = []
    count = 0
    len_list, result, count = len(list), [], 0

    if len_list == 1: return list
    if len_list % n:
        len_sublist = len_list//n + 1
    else:
        len_sublist = len_list//n

    for i in range(n-1):
        result.append(list[count:count+len_sublist])
        count += len_sublist
        if count > len_list:
            break
    if count < len(list):
        result.append([list[count:]])
    return result


def get_col_widths(dataframe):
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]



def rep_txt_wrapper(replist, txtout, obj):
    result = []
    for rep in replist:
        result.append((rep, txtout, obj))
    return result


def replay_file_parser(folderpath):
    filepaths = []
    count = 0
    for path, subdirs, files in walk(folderpath):
        for name in files:
            extension = name[-9:]
            if isfile(join(path, name)) and extension == 'SC2Replay':
                filepaths.append(join(path, name))
                count += 1

    return filepaths, count


def separate_replays_analysis(repl_list, textoutputpath, total_replay_data):
    for rep in repl_list:
        mainprocess(rep, textoutputpath, total_replay_data)


def separate_replaypool(repl_list, textoutputpath, num_of_proc):
    total_replay_data = totalreplaydataclass()

    inputlist = rep_txt_wrapper(repl_list, textoutputpath, total_replay_data)
    pool = multiprocessing.Pool(processes=num_of_proc)
    output = pool.starmap(mainprocess, inputlist)

    for out in output:
        if out is not False:
            total_replay_data.appendtoself(out[0], out[1])

    total_replay_data.create_dataframes()
    total_replay_data.create_excel_file(textoutputpath)
    # total_replay_data.adjust_excel_data(textoutputpath)

    print("All Processes Finished")

    # sublist = splitlist(repl_list, num_of_proc)
    # processlist = []
    # if num_of_proc != len(sublist):
    #     print("Sublist length: ", len(sublist))
    #     print("Number of Processors: ", num_of_proc)
    #
    # for i in range(num_of_proc):
    #     p = multiprocessing.Process(target=separate_replays_analysis, args=(sublist[i], textoutputpath, total_replay_data))
    #     p.start()

    #     processlist.append(p)
    # for i, p in enumerate(processlist):
    #     if i == 0:
    #         p.start()
    #     else:
    #         p.join()

# def check_analyzer_status(p):
#     if p.is_alive(): # Then the process is still running
#         label.config(text = "MP Running")
#         mp_button.config(state = "disabled")
#         not_mp_button.config(state = "disabled")
#         root.after(200, lambda p=p: check_status(p)) # After 200 ms, it will check the status again.
#     else:
#         label.config(text = "MP Not Running")
#         mp_button.config(state = "normal")
#         not_mp_button.config(state = "normal")
#     return

class totalreplaydataclass:
    def __init__(self):
        self.replays_data_human = None
        self.replays_data_zombie = None
        self.replays_data_human_list = []
        self.replaynum = 1
        self.replays_data_zombie_list = []
        self.replaysignatures = set()
        self.duplicatereplaylist = []

    def appendtoself(self, hdata, zdata):
        self.replays_data_human_list.extend(hdata)
        self.replays_data_zombie_list.append(zdata)

    def create_dataframes(self):
        self.replays_data_human = pd.DataFrame.from_records(self.replays_data_human_list, columns=total_df_human_column_list)
        self.replays_data_zombie = pd.DataFrame.from_records(self.replays_data_zombie_list, columns=total_df_zombie_column_list)

    def create_excel_file(self, path):
        self.replays_data_human.to_excel(f'{path}/#Humandata.xlsx', header=True)
        self.replays_data_zombie.to_excel(f'{path}/#Zombiedata.xlsx', header=True)
        # h_writer = pd.ExcelWriter(f'{path}/#Humandata.xlsx', engine='xlsxwriter')
        # self.replays_data_human.to_excel(h_writer, sheet_name='Human Data')
        # h_workbook = h_writer.book
        # h_worksheet = h_writer.sheets['Human Data']
        # h_length = len(self.replays_data_human)+2
        # h_worksheet.add_table(f'A2:AY{h_length}', {'data': self.replays_data_human.values.tolist(),
        #                                            'columns': total_df_human_excel_column_list,
        #                                            'header_row': True})
        #
        # for i, width in enumerate(get_col_widths(self.replays_data_human)):
        #     h_worksheet.set_column(i, i, width)
        #
        # h_writer.close()
        #
        # print(self.replays_data_human)
        #
        # z_writer = pd.ExcelWriter(f'{path}/#Zombiedata.xlsx', engine='xlsxwriter')
        # self.replays_data_zombie.to_excel(z_writer, sheet_name='Zombie Data')
        # z_workbook = z_writer.book
        # z_worksheet = z_writer.sheets['Zombie Data']
        # z_length = len(self.replays_data_zombie)+2
        # z_worksheet.add_table(f'A2:AZ{z_length}', {'data': self.replays_data_zombie.values.tolist(),
        #                                            'columns': total_df_zombie_excel_column_list,
        #                                            'header_row': True})
        #
        # for i, width in enumerate(get_col_widths(self.replays_data_zombie)):
        #     z_worksheet.set_column(i, i, width)
        #
        # z_writer.close()
        #
        # print(self.replays_data_zombie)

    # def adjust_excel_data(self, path):
    #     hwb = openpyxl.load_workbook(filename=f'{path}/#Humandata.xlsx')
    #     hsheet_ranges = hwb['range names']


        # h_workbook = xlsxwriter.Workbook(f'{path}/#Humandata.xlsx')
        # print(h_workbook.worksheets())
        # h_worksheet = h_workbook.worksheets()[0]
        # h_length = len(self.replays_data_human) + 1
        # h_worksheet.add_table(f'B1:AZ{h_length}', {'header_row': True})
        #
        # for i, width in enumerate(get_col_widths(self.replays_data_human)):
        #     h_worksheet.set_column(i, i, width)
        #
        # h_workbook.close()
        #
        # z_workbook = xlsxwriter.Workbook(f'{path}/#Zombiedata.xlsx')
        # z_worksheet = z_workbook.worksheets()[0]
        # z_length = len(self.replays_data_zombie) + 1
        # z_worksheet.add_table(f'B1:BA{z_length}', {'header_row': True})
        #
        # for i, width in enumerate(get_col_widths(self.replays_data_zombie)):
        #     z_worksheet.set_column(i, i, width)
        #
        # z_workbook.close()



# https://xlsxwriter.readthedocs.io/working_with_pandas.html
# https://stackoverflow.com/questions/52052184/how-to-use-xlsxwriter-add-table-method-with-a-dataframe


class maininfoclass:
    def __init__(self):
        self.replayfolderpath = ""
        self.textfolderpath = ""
        self.replayfilepaths = []
        self.replaycount = 0
        self.quickanalysis = False
        self.quickcheck = "mainclass load complete!"
        self.version = ""

        self.numberofprocesses = 3

    # Depreciated in favor of GUI-local functions (no real reason, these still work)
    def get_replays_from_folder(self):
        self.replayfilepaths, self.replaycount = replay_file_parser(self.replayfolderpath)

    # Depreciated in favor of local functions (they cause multiple instances of GUI to pop up)
    # def send_replays_to_self(self):
    #     if len(self.replayfilepaths) > 1:
    #         # self.replaypool_loop(self.replayfilepaths)
    #         separate_replaypool(self.replayfilepaths, self.numberofprocesses)
    #     else:
    #         print("No replays detected!")
    #         pass
    #
    # def analyze_replays(self, replaypaths):          # pass multiple analyze_replays w different lists to replayque_loop
    #     print("analysis started!")
    #     for replaypath in replaypaths:
    #         mainprocess(replaypath, self.quickanalysis)
    #
    # def replaypool_loop(self, list_of_replays):
    #     pool = multiprocessing.Pool(self.numberofprocesses)
    #     pool.map(separate_replays_analysis, list_of_replays)



# https://docs.python.org/3/library/multiprocessing.html
# https://towardsdatascience.com/multiprocessing-in-python-9d498b1029ca
# https://superfastpython.com/multiprocessing-in-python/#How_to_Run_a_Function_In_a_Process

# https://superfastpython.com/multiprocessing-semaphore-in-python/
# check out the above if data leak happens
